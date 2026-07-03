"""Dataset loaders for the new guide-format datasets (xlsx, csv, txt).

Each loader reads the new dataset format and returns data structures compatible
with the existing pipeline consumers, so the rest of the code doesn't need to
know which format it's working with.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

from prior_auth.utils.paths import DATASET_ROOT


# ---------------------------------------------------------------------------
# Doctor Notes (xlsx → list of dicts matching legacy JSON case format)
# ---------------------------------------------------------------------------

def load_doctor_notes_xlsx(
    path: Path | None = None,
) -> list[dict[str, Any]]:
    """Load doctor notes from the new xlsx format.

    Returns a list of dicts with keys: case_id, specialty, note_text
    (matching the legacy JSON case format used by PriorAuthWorkflow).
    """
    path = path or (DATASET_ROOT / "doctor_notes" / "doctor_notes.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cases: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        cases.append({
            "case_id": record["case_id"],
            "specialty": record.get("specialty", "unknown"),
            "note_text": record.get("doctor_note", ""),
        })
    wb.close()
    return cases


# ---------------------------------------------------------------------------
# ICD-10 Master CSV → list of dicts matching legacy KG JSON record format
# ---------------------------------------------------------------------------

_LATERALITY_PATTERNS = {
    "left": re.compile(r"\bleft\b", re.IGNORECASE),
    "right": re.compile(r"\bright\b", re.IGNORECASE),
    "bilateral": re.compile(r"\bbilateral\b", re.IGNORECASE),
}


def _derive_laterality(description: str) -> str:
    """Infer laterality from the diagnosis description."""
    for lat, pat in _LATERALITY_PATTERNS.items():
        if pat.search(description):
            return lat
    return "not_applicable"


def _derive_category(icd_code: str) -> str:
    """Derive the ICD-10 category (letter + digits before the dot)."""
    return icd_code.split(".")[0] if "." in icd_code else icd_code


def _derive_keywords(description: str) -> list[str]:
    """Tokenize the diagnosis description into keywords, stripping laterality
    words and very short/generic tokens.  This is not perfect, but gives the
    keyword-matcher something to work with without hand-authoring keyword lists.
    """
    stopwords = {
        "a", "an", "the", "of", "with", "for", "and", "or", "in", "on", "to",
        "is", "are", "primary", "secondary", "unspecified",
    }
    words = re.findall(r"[a-z0-9']+", description.lower())
    keywords = [w for w in words if w not in stopwords and len(w) > 2]

    # Also add the laterality-stripped full description as a phrase keyword
    stripped = re.sub(r"\b(left|right|bilateral)\b", "", description, flags=re.IGNORECASE).strip()
    stripped = re.sub(r"\s{2,}", " ", stripped)
    if stripped and stripped.lower() != description.lower():
        keywords.append(stripped.lower())

    return keywords


def _derive_specialty(description: str, icd_code: str) -> str:
    """Best-effort specialty inference from diagnosis text and code prefix."""
    desc_lower = description.lower()

    specialty_hints = {
        "orthopedics": ["knee", "hip", "joint", "osteoarthritis", "fracture", "spine", "rotator", "acl", "meniscus"],
        "cardiology": ["coronary", "heart", "cardiac", "angina", "artery disease", "aortic", "atrial"],
        "neurology": ["parkinson", "epilepsy", "migraine", "multiple sclerosis", "brain stimulation", "seizure"],
        "oncology": ["cancer", "malignant", "neoplasm", "tumor", "carcinoma", "lymphoma", "leukemia"],
        "gastroenterology": ["crohn", "colitis", "liver", "gastric", "bariatric", "biliary", "pancrea"],
        "rare_disease": ["gaucher", "fabry", "pompe", "wilson", "cystic fibrosis", "huntington"],
    }

    for specialty, hints in specialty_hints.items():
        if any(h in desc_lower for h in hints):
            return specialty

    # Code-prefix fallback
    code_letter = icd_code[0].upper() if icd_code else ""
    code_map = {"M": "orthopedics", "I": "cardiology", "G": "neurology", "C": "oncology", "K": "gastroenterology"}
    return code_map.get(code_letter, "general")


def load_icd_csv(
    path: Path | None = None,
) -> list[dict[str, Any]]:
    """Load ICD codes from the new CSV format and derive the fields the
    ICD10KnowledgeGraph expects (category, keywords, laterality, specialty).

    Returns a list of dicts matching the legacy JSON record schema.
    """
    path = path or (DATASET_ROOT / "icd" / "icd10_master.csv")
    records: list[dict[str, Any]] = []

    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row["icd10_code"]
            description = row["diagnosis"]
            is_rare = row.get("rare_disease", "No").strip().lower() in ("yes", "true", "1")

            records.append({
                "code": code,
                "description": description,
                "category": _derive_category(code),
                "specialty": _derive_specialty(description, code),
                "laterality": _derive_laterality(description),
                "keywords": _derive_keywords(description),
                "rare_disease": is_rare,
            })

    return records


# ---------------------------------------------------------------------------
# Policies (prose .txt + criteria .json sidecars)
# ---------------------------------------------------------------------------

def load_policies_from_dir(
    policy_dir: Path | None = None,
) -> list[dict[str, Any]]:
    """Load policies from the new format: prose .txt files with optional
    .criteria.json sidecars.

    Returns a list of dicts compatible with the PolicyRAG consumer.
    """
    policy_dir = policy_dir or (DATASET_ROOT / "policies")
    policies: list[dict[str, Any]] = []

    for txt_path in sorted(policy_dir.glob("*_policy.txt")):
        specialty = txt_path.stem.replace("_policy", "")
        policy_text = txt_path.read_text(encoding="utf-8").strip()

        # Look for a criteria sidecar
        criteria_path = txt_path.with_suffix("").with_suffix(".criteria.json")
        criteria: list[dict] = []
        if criteria_path.exists():
            with open(criteria_path, "r", encoding="utf-8") as f:
                criteria = json.load(f)

        # Derive procedure/diagnosis keywords from the policy text
        procedure_keywords = _extract_procedure_keywords(policy_text)
        diagnosis_keywords = _extract_diagnosis_keywords(policy_text)

        policies.append({
            "policy_id": f"POL-{specialty.upper()}-NEW-001",
            "title": f"{specialty.title()} Medical Necessity Policy",
            "specialty": specialty,
            "procedure_keywords": procedure_keywords,
            "diagnosis_keywords": diagnosis_keywords,
            "criteria": criteria,
            "policy_text": policy_text,
        })

    return policies


def _extract_procedure_keywords(text: str) -> list[str]:
    """Extract procedure names mentioned in policy text."""
    procedures = []
    # Look for known procedure patterns
    patterns = [
        r"\b(PCI|CABG|DBS|TAVR|ERCP)\b",
        r"\b(total knee replacement|total hip replacement|knee arthroplasty)\b",
        r"\b(chemotherapy|radiation therapy|biologic therapy)\b",
        r"\b(catheter ablation|pacemaker implantation)\b",
        r"\b(deep brain stimulation)\b",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            procedures.append(m.group(0).lower())
    return list(set(procedures)) if procedures else [text.split()[0].lower()]


def _extract_diagnosis_keywords(text: str) -> list[str]:
    """Extract diagnosis-related keywords from policy text."""
    keywords = []
    patterns = [
        r"\b(CAD|coronary artery disease)\b",
        r"\b(osteoarthritis|OA)\b",
        r"\b(Crohn disease|ulcerative colitis|IBD)\b",
        r"\b(Parkinson disease|parkinsonism)\b",
        r"\b(malignancy|cancer|neoplasm)\b",
        r"\b(breast cancer|lung cancer)\b",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            keywords.append(m.group(0).lower())
    return list(set(keywords)) if keywords else []


# ---------------------------------------------------------------------------
# Form Template (new flat JSON → legacy multi-template format)
# ---------------------------------------------------------------------------

def load_form_template(
    path: Path | None = None,
) -> list[dict[str, Any]]:
    """Load the new flat form template and wrap it in the list-of-templates
    format the FormFillerAgent expects.
    """
    path = path or (DATASET_ROOT / "forms" / "prior_auth_template.json")
    with open(path, "r", encoding="utf-8") as f:
        template = json.load(f)

    # Wrap in the format expected by the form filler
    return [{
        "template_id": "GUIDE-PA-TEMPLATE-001",
        "title": "Prior Authorization Form",
        "fields": list(template.keys()),
        "template": template,
    }]


# ---------------------------------------------------------------------------
# Gold Standard Evaluation (xlsx)
# ---------------------------------------------------------------------------

def load_gold_standard_xlsx(
    path: Path | None = None,
) -> list[dict[str, Any]]:
    """Load gold standard evaluation data from xlsx.

    Returns a list of dicts with keys: case_id, diagnosis, expected_extractor_json.
    """
    path = path or (DATASET_ROOT / "evaluation" / "gold_standard_evaluation.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Parse the expected JSON string into a dict
        expected_json_str = record.get("expected_extractor_json", "{}")
        try:
            expected = json.loads(expected_json_str)
        except (json.JSONDecodeError, TypeError):
            expected = {}
        records.append({
            "case_id": record["case_id"],
            "diagnosis": record.get("diagnosis", ""),
            "expected": expected,
        })
    wb.close()
    return records


# ---------------------------------------------------------------------------
# Edge Cases (xlsx)
# ---------------------------------------------------------------------------

def load_edge_cases_xlsx(
    path: Path | None = None,
) -> list[dict[str, Any]]:
    """Load edge cases from xlsx.

    Returns a list of dicts with keys: case_id, edge_case, description.
    """
    path = path or (DATASET_ROOT / "edge_cases" / "edge_cases.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        records.append({
            "case_id": record.get("case_id", ""),
            "edge_case": record.get("edge_case", ""),
            "description": record.get("description", ""),
        })
    wb.close()
    return records
